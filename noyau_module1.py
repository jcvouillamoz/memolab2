# -*- coding: utf-8 -*-
"""
Created on Sun Feb 28 20:54:17 2021

Noyau - Module 1

Ce module est constitué de fonctions autonomes et donc sans structure
de classes. Les données sont échangées par les paramètres en entrées et
les valeurs retournées en sortie.

Ces fonctions sont donc aussi exploitables depuis tous les autres modules
sans rique de bouclages logique circulaire.

INDEX DES FONCTIONS
ouSuisJe():
    cette fonction retourne le chemin complet du script en cours d'exécution
chargeParametresGeneraux(nomFichierParametres):
    Création et renvoi paraGen[] contenant les données de params_generaux.xlsx
xlsx2List(nomFichierXlsx,listXlsx,paraGen):
    Création et renvoi listXlsx[] contenant les données de nomFichierXlsx.xlsx
list2Xlsx(listXlsx,nomFichierXlsx,paraGen):
    Transfert du contenu d'une liste xlsx vers un fichier Excel 
message(titreMessage, listMessages):
    retourne le message correspondant à son titre
extraitNomFichier(nomComplet):
    Extraction du nom de fichier d'une chaine complète


@author: JCV
"""
# Récupération chemin actuel du script
def ouSuisJe():
    """
    cette fonction retourne le chemin complet du script en cours d'exécution'

    Returns
    -------
    chemin : str
        Chemin complet du script en cours d'exécution.
        Exemple : c:/user/jcv/memolab/

    """
    import os
    absFilePath = os.path.abspath(__file__)
    chemin, nomScript = os.path.split(absFilePath)
    chemin = chemin.replace("\\","/")    # le \ est doublé mais compte simple
    return chemin


def chargeParametresGeneraux(nomFichierParametres):
    """
    Charge les paramètres contenus dans le fichier xlsx appelé dans le
    dictionnaire paraGen retourné au script appelant

    Parameters
    ----------
    nomFichierParametres : str
        Nom du fichier excel xlsx contenant la liste des paramètres selon
        le modèle fourni avec l'appli. Ex: params_generaux.xlsx

    Returns
    -------
    paraGen : dict
        Un terme par paramètre.
        L'appel d'un paramètre se fait ainsi :
            <valeurParamètre> = paraGen["<nomParamètre>"]

    """
    paraGen = {}
    # Accès au fichier excel
    import openpyxl
    # chemin du script courant
    chemin = ouSuisJe()
    # print(chemin)
    # objet workbook
    wb = openpyxl.load_workbook(chemin + "/" + nomFichierParametres)
    # objet worksheet
    ws = wb.active
    for ligne in range(1, 100):
        adrCelluleNom = "B" + str(ligne)
        adrCelluleVal = "C" + str(ligne)
        nomParametre = ws[adrCelluleNom].value
        valeurParametre = ws[adrCelluleVal].value
        if valeurParametre == None:
            break
        # ajout parametre au dictionnaire
        paraGen.update({nomParametre : valeurParametre})
        # essai création variable
        # exec(nomParametre + " = valeurParametre")
    wb.close()
    # print("JeanClaude =", JeanClaude)
    return paraGen

def xlsx2List(nomFichierXlsx,listXlsx,paraGen,cheminRelXlsx="/data"):
    """ Transfert du contenu de la feuille xlsx vers la liste 
    
    Paramètres:
        nomFichierXlsx : str : exemple: user.xlsx
        listXlsx : list
        paraGen : Dictionnaire des paramètres généraux
    
    Le nom complet avec chemin est formé ici au moyen de la
    fonction OuSuisJe() et du paraGen du nom du sous-pépertoire
    relatif contenant le fichier à ouvrir. (ex: data/)
    Le sous-répertoire de paraGen[] ne peut être vide. Les fichiers xlsx
    ouverts ici doivent être dans un autre dossier que la racinde de Memolab.
    
    Le nom de liste passé en paramètre doit correspondre à une
    liste existante, dont le contenu sera substitué par celui
    du fichier excel, ou alors initialisé au préalable par 
    l'appelant.
    
    Cette fonction retourne la liste listXlsx
    """
    # Etablissement du nom de fichier xlsx complet
    racineChemin = ouSuisJe()
    racineCheminSRep = racineChemin + cheminRelXlsx + "/"
    nomXlsxComplet = racineCheminSRep + nomFichierXlsx
    
    # accès au fichier excel et à sa feuille active
    import openpyxl
    wb = openpyxl.load_workbook(nomXlsxComplet,data_only=True)
    ws = wb.active
    
    # chargement de la feuille dans la liste
    for i, row in enumerate(ws.rows):
            listXlsx.append([])
            for cell in row:
                listXlsx[i].append(cell.value)
    
    # cloture accès fichier xlsx
    wb.close()

    return listXlsx    
    
""" # ################# code test xlsx2List #############
# charge paraGen
nomFichierParametres = "params_generaux.xlsx"
paraGen = chargeParametresGeneraux(nomFichierParametres)
nomFichierXlsx = "classeurExcel1.xlsx"
listXlsx=[]
listXlsx = xlsx2List(nomFichierXlsx,listXlsx,paraGen)
if paraGen["imprimeOK"]:
    print(listXlsx)
""" # ###################################################    
    

def list2Xlsx(listXlsx,nomFichierXlsx,paraGen,cheminRelXlsx="/data"):
    """ Transfert du contenu d'une liste xlsx vers un fichier Excel 
    
    Paramètres:
        listXlsx : list
        nomFichierXlsx : str : exemple: user.xlsx
        paraGen : Dictionnaire des paramètres généraux
    
    Le nom complet avec chemin est formé ici au moyen de la
    fonction OuSuisJe() et du paraGen du nom du sous-pépertoire
    relatif contenant le fichier à ouvrir. (ex: /data)
    
    Le nom de liste passé en paramètre doit correspondre à une
    liste existante. Sa structure sera compatible avec celle de la
    feuille excel destinataire. Il s'agira typiquement d'une feuille
    excel qui a été chargée dans la liste, puis cette liste a
    été modifiée par le script. Et on actualise ici ces changements
    sur la feuille excel d'origine.
    
    Cette fonction ne retourne rien
    """    
    # Etablissement du nom de fichier xlsx complet
    racineChemin = ouSuisJe()
    racineCheminSRep = racineChemin + cheminRelXlsx + "/"
    nomXlsxComplet = racineCheminSRep + nomFichierXlsx
    # if paraGen["imprimeOK"]:
    #     print("xlsx2List : nomXlsxComplet : {}".format(nomXlsxComplet))
        
    # accès au fichier excel et à sa feuille active
    import openpyxl
    wb = openpyxl.load_workbook(nomXlsxComplet,data_only=True)
    ws = wb.active
    
    # chargement de la liste dans la feuille

    """ Evaluation nbre de lignes et de colonnes à transférer
    Partons du principe que nous traitons une liste 2D rectangulaire
    dont toutes les cellules de la ligne 0 contiennent quelque chose.
    Il en résultera que le nombre de lignes de la liste sera obtenue
    par la fonction len(list) et que le nombre de colonnes de la liste
    sera obtenu par la fonction len(list[0])
    """
    nbreLignes = len(listXlsx)
    nbreColonnes = len(listXlsx[0])
    print("Module 1 : nbreLignes : ", nbreLignes)
    print("Module 1 : nbreColonnes : ", nbreColonnes)
    

    # boucle écriture
    for ligne in range(1,nbreLignes+1):
        for colonne in range(1,nbreColonnes+1):
            ws.cell(row=ligne, column=colonne, value=listXlsx[ligne-1][colonne-1])
    
    # Sauvegarde de la feuille excel modifiée
    wb.save(nomXlsxComplet)
    
 
"""# ##### code test xlsx2List et list2Xlsx #############
# charge paraGen
nomFichierParametres = "params_generaux.xlsx"
paraGen = chargeParametresGeneraux(nomFichierParametres)
nomFichierXlsx = "essai.xlsx"
listXlsx=[]
# charge liste 
listXlsx = xlsx2List(nomFichierXlsx,listXlsx,paraGen)
if paraGen["imprimeOK"]:
    print(listXlsx)
    
# modification d'un élément de la liste
listXlsx[0][0] = "Origine"

# sauve liste dans fichier Excel
list2Xlsx(listXlsx,nomFichierXlsx,paraGen)
"""# ###################################################  

def message(titreMessage, listMessages):
    """ Cette fonction retourne le message correspondant à son titre.
    
    Les messages sont enregistrés dans data/messages.xlsx. Ils sont
    chargés dans la listMessages[], en entête du script main.

    Parameters
    ----------
    titreMessage : str
        La 2e colonne de la liste contient les titres de chaque messages
        situés, eux dans la 3e colonne.

    Returns
    -------
    message : str

    """
    for element in enumerate(listMessages):
        # element est un tuple de la ligne en cours
        #    print("element : {} \n Type(element) : {} ".format(element, type(element)))
        # titre est une liste de la ligne en cours
        listElement = element[1]
        # listElement est une liste de la ligne en cours
        #    print("listElement : {} \n Type(listElement) : {} ".format(listElement, type(listElement)))
        titre = listElement[1]
        # titre est le titre de la ligne en cours
        #    print("titre : {} \n Type(titre) : {} ".format(titre, type(titre)))
        if titre == titreMessage:
            #    print("Trouvé !")
            # cherche \n dans le message
            message = str(listElement[2])
            print(message)
            pos = message.find("\\n")
            if pos > 0:
                print("Remplacements de \\n par des \n")
                message = message.replace("\\n","\n")
                message = message.replace("\\","")
                print(message)
            print(pos)
            return message
            break
    return "Message introuvable"

def annonce(titre, message, aspect=200, taillePoliceCar=12, editeOK=False, couleur="#ffffff"):
    import tkinter
    import math
    
    # Déclarations
    largeurCaractereEnPix = 9   # en taille standard
    hauteurCaractereEnPix = 10  # en taille standard
    
    # calculs taille de la fenêtre
    nbCaracteres = len(message) ** 1.15
    largeurEnCaracteres = max(int(math.sqrt(aspect*nbCaracteres/100)),15)
    hauteurEnLignes = max(int(nbCaracteres / largeurEnCaracteres) + 2, 20)
    facteurTaillePoliceCaractere = taillePoliceCar / 10
    largeurEnPix = int(largeurEnCaracteres * largeurCaractereEnPix * facteurTaillePoliceCaractere)
    hauteurEnPix = int(hauteurEnLignes * hauteurCaractereEnPix * facteurTaillePoliceCaractere)

    # Création fenetre
    fenetre = tkinter.Tk()
    fenetre.configure(width=largeurEnPix,height=hauteurEnPix)
    fenetre.pack_propagate(False)   # gèle la taille de la fenêtre à sa consigne
    fenetre.title(titre)

    # Centrage fenêtre
    largeurEcran = int(fenetre.winfo_screenwidth())
    hauteurEcran = int(fenetre.winfo_screenheight())
    largeurFenetre = largeurEnPix
    hauteurFenetre = hauteurEnPix
    positionX = largeurEcran // 2 - largeurFenetre // 2
    positionY = hauteurEcran // 2 - hauteurFenetre // 2
    paramGeometry = "{}x{}+{}+{}".format(largeurFenetre,hauteurFenetre,positionX,positionY)
    fenetre.geometry(paramGeometry)


    # Création objet scrollbar
    scrollbar = tkinter.Scrollbar(fenetre)
    
    # Création widget Text avec sa scrollbar verticale associée
    texte = tkinter.Text(fenetre, yscrollcommand=scrollbar.set)
    scrollbar.config(command=texte.yview)
    scrollbar.pack(side=tkinter.RIGHT, fill=tkinter.Y)
    
    # Configuration du visuel du text
    texte.configure(font=("Helvetica",taillePoliceCar))
    texte.configure(padx=20, pady=20)
    texte.configure(wrap="word")
    texte.configure(bg=couleur)
   
    # Droit modifier le contenu du texte
    if not editeOK:
        texte.bind("<Key>", lambda a: "break")
    
    # Assignation du titre et du message à Text
    """ 
    Le premier paramètre 1.0 indique la position d'insertion du texte ou autre
    objet inséré. 1 désigne le numéro de ligne et 0 le numéro de colonne.
    Alternativemement, pour ajouter à la fin du contenu de Text, on peut
    utiliser tkinter.END
    """
    texte.insert(1.0, titre)
    texte.insert(tkinter.END,"\n\n"+message)

    # mise en gras du titre
    texte.tag_add("debutEnGras",1.0, "1."+str(len(titre)))
    texte.tag_config("debutEnGras", font=("Helvetica",taillePoliceCar, "bold"))
    texte.update()
    
    # Frame des boutons
    frameBoutons = tkinter.Frame(fenetre)
    frameBoutons.pack(side=tkinter.BOTTOM)
    
    # Boutons
    boutonCancel = tkinter.Button(frameBoutons, text="Concel", command = fenetre.destroy)
    boutonCancel.grid(row=0, column=0)
    boutonQuitter = tkinter.Button(frameBoutons, text="OK", command = fenetre.destroy)
    boutonQuitter.grid(row=0, column=1)
    
    
    # pack du Text avec fill
    """
    Le paramètre fill indique la direction d'expansion du widget. Les 
    valeurs possibles sont: 
        tkinter.X en largeur
        tkinter.Y en hauteur
        tkinter.BOTH enlargeur et hauteur
    """    
    texte.pack(expand=tkinter.YES, fill=tkinter.X)
    
    fenetre.mainloop()

def __afficheQuestionSaisiReponse(question, largeurQuestion=640):
    """
    Cette fonction est strictement au service de la fonction questionne() ci-dessous.
    Elle affiche une fenêtre popup contenuant une question (text seulement).
    Elle ouvre un champ de saisie d'un text multi lignes pour la réponse.
    Cette réponse est ensuite écrite dans un fichier temporaire temp.txt afin
    d'être récupérée par la fonction appelante. Ainsi est contourné le défi
    de persistance des données aussi éphémères que les widgets qui les utilisent.

    Parameters
    ----------
    question : str
        texte de la question posée
    largeurQuestion : str, optional
        The default is 640.
        Largeur en pixels du widget Message qui affiche la question et détermine
        également la largeur de la fen^tre et du widget Text de la réponse 
        situé dessous.

    Returns
    -------
    None.

    """
    import time
    import tkinter
    fenetre = tkinter.Tk()
    
    def __afficheQuestion(fenetre,question):
        messageQuestion = tkinter.Message(fenetre,
                                          text=question,
                                          width=largeurQuestion)
        messageQuestion.grid(row=0, column=0)
        
    def __SaisiReponse(fenetre):
        
        def __getText():
            reponse=textReponse.get(1.0, "end")
            with open("temp.txt", "w") as filout:
                for ligne in reponse:
                    filout.write(ligne)
            
            """
            dbmFile['Reponse'] = reponse
            dbmFile.close()
            """
            
            textReponse.destroy()
            time.sleep(1)
            fenetre.destroy() 
        
        textReponse = tkinter.Text(fenetre, wrap="word", pady=20, padx=20)
        textReponse.grid(row=1,column=0)
        buttonSaveAndQuit = tkinter.Button(fenetre,text="Save & Quit", command=__getText)
        buttonSaveAndQuit.grid(row=2, column=0)
    
    __afficheQuestion(fenetre,question)
    __SaisiReponse(fenetre)
    
    fenetre.mainloop()
    
def questionne(question):
    __afficheQuestionSaisiReponse(question)

    reponse = ""
    with open("temp.txt", "r") as filin:
        ligne = filin.readline()
        while ligne != "":
            reponse += ligne
            ligne = filin.readline()

    # print(reponse)
    return reponse

""" # programme de test et démonstration
if __name__ == "__main__":
    
    reponse = questionne("Qui es-tu ?")
    print(reponse)
"""  


    
def extraitNomFichier(nomComplet):
    """
    Extraction du nom de fichier d'une chaine complète

    Parameters
    ----------
    nomComplet : str
        nom complet du fichier depuis le disque avec path 

    Returns
    -------
    nom du fichier sans le chemin

    """
    # Repérer la position du dernier /
    positionAvantSlach = nomComplet.rfind("/")
    # print(positionAvantSlach)
    
    nomFichier = ""
    compteur = 0
    for car in nomComplet:
        if compteur > positionAvantSlach:
            nomFichier += str(car)
        compteur += 1
    # print(nomFichier)
    
    return nomFichier

def creationListTest(nbLignes=20, nbColonnes=20, largeurCellules=10):
    """
    génération d'une liste pour tester, en vue de traiter n'importe quelle liste
    en entrée par la suite
    """
    # nbLignes = 60
    # nbColonnes = 20
    nbCarLargCel = largeurCellules
    listTest = []
    # boucle des lignes
    for posLigne in range(nbLignes):
        # ajout d'une ligne
        listTest.append([])
        # boucle des colonnes
        for posColonne in range(nbColonnes):
            contenuBrutCell = str(posLigne) + "/" + str(posColonne) + "-" + "X"*nbCarLargCel
            contenuCell = contenuBrutCell[:nbCarLargCel]
            listTest[posLigne].append(contenuCell)
    # print(listTest)
    return listTest


def afficheListEnTable(listAAfficher):
    import tkinter
    fenetre = tkinter.Tk()
        
    # Création canva dans fenetre
    canvaFrame = tkinter.Canvas(fenetre)
    # création frame dans canva
    frameTable = tkinter.Frame(canvaFrame)
    # créations des scrollbars dans fenetre
    scrollbarHorizontale = tkinter.Scrollbar(fenetre)
    scrollbarVerticale = tkinter.Scrollbar(fenetre)
    
    def updateScrollRegion():
        canvaFrame.update_idletasks()
        canvaFrame.config(scrollregion=frameTable.bbox())
    
    # réglages du Canvas, du Frame, et des scrollbars pour le scrolling
    def __createScrollableContainer():
        # rattachement au canva des commandes des scrollbars
    	canvaFrame.config(xscrollcommand=scrollbarHorizontale.set,
                       yscrollcommand=scrollbarVerticale.set, 
                       highlightthickness=0)
        # config orientations des scrollbars et de leur pilotage 
        # par les vues de canvaFrame
    	scrollbarHorizontale.config(orient=tkinter.HORIZONTAL, 
                                 command=canvaFrame.xview)
    	scrollbarVerticale.config(orient=tkinter.VERTICAL, 
                               command=canvaFrame.yview)
        # affichage des scrollbars
    	scrollbarHorizontale.pack(fill=tkinter.X, side=tkinter.BOTTOM, 
                               expand=tkinter.FALSE)
    	scrollbarVerticale.pack(fill=tkinter.Y, side=tkinter.RIGHT, 
                             expand=tkinter.FALSE)
        # affichage du caneva
    	canvaFrame.pack(fill=tkinter.BOTH, side=tkinter.LEFT, expand=tkinter.TRUE)
        # affichage du frame dans le canva
    	canvaFrame.create_window(0, 0, window=frameTable, anchor=tkinter.NW)
    
    def __afficheTableList(listAAfficher):
        """
        global i
        tkinter.Label(frameTable, text="Hello World").grid(row=i, column=i)
        i+=1
        """
        # print("Affiche liste tabulaire")
        # print(listAAfficher)
        nbLignes = len(listAAfficher)
        nbColonnes = len(listAAfficher[0])
        # boucle de création des labels
        for posLigne in range(nbLignes):
            for posColonne in range(nbColonnes):
                contenuCell = listAAfficher[posLigne][posColonne]
                # print(contenuCell)
                tkinter.Label(frameTable, text=contenuCell).grid(row=posLigne, column=posColonne)
    
        # Mise à jour après l'ajout dynamique de nouveau widget
        updateScrollRegion()
    
    __createScrollableContainer()
    # addNewLabel()
    # listTest = creationListTest()
    __afficheTableList(listAAfficher)
    
    fenetre.mainloop()





# Création nom de fichier à partir d'un nom d'utilisateur quelconque
def creaNomFichier(nomUtilisateur):
    # Caractères utilisables dans nomUtilisateur
    carUtilisables="abcdefghijklmnopqrstuvwxyz1234567890"
    # boucle filtrage nomUtilisateur
    nomFichier = ""
    for car in nomUtilisateur:
        if car.lower() in carUtilisables:
            nomFichier += car.lower()
    nomFichier += ".xlsx"

    return nomFichier

      
####################### Comparaisons de chaînes de caractèers #############

def CompareIdentique(mot1, mot2, EspacesExacts = False):
    """
    Compare deux chaines avec tolérance de différence d'espacements
    entre les mots

    Parameters
    ----------
    mot1 : TYPE Str
        Premier mot à comparer
    mot2 : TYPE Str
        Second mot à comparer
    EspacesExacts : Type Bool
        Tient compte ou non du nombre d'espaces entre les mots et au début
        et fin des chaînes

    Returns
    -------
    TYPE Bool
        Résultat de la comparaison

    """
    if not(EspacesExacts):
        mot1 = mot1.split()
        mot2 = mot2.split()
    return mot1==mot2

"""
# test fonction
mot1 = "Il était une fois une  histoire dont tout reste à inventer"
mot2 = "Il était une fois une histoire dont tout reste à inventer"
print(CompareIdentique(mot1, mot2, EspacesExacts = True))
"""


def CompareIdentiqueMinuscule(mot1, mot2,EspacesExacts = False):
    """
    Comparaison de deux chaînes converties en minuscules et optionnellement
    sans tenir compte du nombre d'espaces entre les mots

    Parameters
    ----------
    mot1 : Str
        Premier mot à comparer
    mot2 : Str
        Second mot à comparer
    EspacesExacts : Bool, optional
        Prise en compte éventuelle du nombre d'espaces entre les mots.
        The default is False.

    TYPE Bool
        Résultat de la comparaison

    """
    mot1 = mot1.casefold()
    mot2 = mot2.casefold()
    return CompareIdentique(mot1, mot2, EspacesExacts)

"""
# test fonction
mot1 = "Il était une fois une Histoire  dont tout reste à inventer"
mot2 = "Il était une fois une histoire dont tout reste à inventer"
print(CompareIdentiqueMinuscule(mot1, mot2, EspacesExacts = True))
"""


def CompareDesordreExhaustif(chaine1, chaine2, casseExacte=False, nbMotsExacts=False):
    """
    Vérifie que tous les mots contenus dans une chaine soient présents dans
    l'autre et réciproquement

    Parameters
    ----------
    chaine1 : Str
        Première chaîne comparée
    chaine2 : Str
        Seconde chaîne comparée
    casseExacte : Bool, optional
        Prise en compte de la casse . The default is False.
    nbMotsExacts : bool, optional
        Correspondance du nombre de chaque mot. The default is False.

    Returns
    -------
    Trouve : bool
        Résultat de la comparaison

    """
    # Mise optionnelle en minuscules
    if not(casseExacte):
        chaine1 = chaine1.casefold()
        chaine2 = chaine2.casefold()
    # Est-ce que tous les mots de chaine2 se retrouvent une ou plusieurs fois
    # dans chaîne1 ?
    # Extraction des mots de chaine2
    dicoChaine2 = chaine2.split()
    # Recherche de chaque mot de chaine2 dans chaine1
    Trouve = True
    for mot in dicoChaine2:
        if not(mot in chaine1):
            Trouve = False
    # Est-ce que tous les mots de chaine1 se retrouvent une ou plusieurs fois
    # dans chaîne2 ?
    # Extraction des mots de chaine1
    dicoChaine1 = chaine1.split()
    # Recherche de chaque mot de chaine2 dans chaine1
    for mot in dicoChaine1:
        if not(mot in chaine2):
            Trouve = False
    # Vérif optionnelle du même nombre de mots identiques des deux côtés
    if Trouve and nbMotsExacts:
        for mot in dicoChaine1:
            nbMotInChaine1 = chaine1.count(mot)
            nbMotInChaine2 = chaine2.count(mot)
            if nbMotInChaine1 != nbMotInChaine2:
                Trouve = False

    return Trouve

"""
# test fonction
mot1 = "Il était une fois une histoire dont tout reste à inventer"
mot2 = "Il était une fois une histoire dont tout reste à inventer"
print(CompareDesordreExhaustif(mot1, mot2, casseExacte = True, nbMotsExacts = True))
"""


def CompareAuMoinsUnMotCommun(chaine1, chaine2, casseExacte=False):
    """
    Véréfie si dans chaîne1 et chaine2 se trouve au moins un mot commun,
    tenant compte optionnellement compte de la casse, et ne tenant pas
    compte des espaces entre les mots

    Parameters
    ----------
    chaine1 : Str
        Première chaine de caractères
    chaine2 : Str
        Seconde chaîne de caractères
    casseExacte : bool, optional
        prise en compte de la casse. The default is False.

    Returns
    -------
    Trouve : bool
        Renvoie True si un mot au moins a été trouvé.

    """
     # Mise optionnelle en minuscules
    if not(casseExacte):
        chaine1 = chaine1.casefold()
        chaine2 = chaine2.casefold()
    # Est-ce que un mot de chaine2 se retrouve dans chaine1 ?
    # Extraction des mots de chaine2
    dicoChaine2 = chaine2.split()
    # Recherche de chaque mot de chaine2 dans chaine1
    Trouve = False
    for mot in dicoChaine2:
        if mot in chaine1:
            Trouve = True

    return Trouve


# =============================== TESTS LOCAUX ========================

if __name__ == '__main__':
    
    # test questionne()
    
    titre = "Question"
    question = "En quelle années est-ce que Jean-Claude Vouillamoz et Suzanne se sont-ils marés et où ?"
    
    reponseRecue = questionne(titre, question)
    
    print(reponseRecue)
    
    
    
    
    